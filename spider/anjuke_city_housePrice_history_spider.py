import requests
from bs4 import BeautifulSoup
import re
import time
import random
import json
import os
import sys
from openpyxl import Workbook


city_name: str = 'beijing'
start_year: int = 2011
end_year: int = 2020
script_path: str = os.path.dirname(os.path.realpath(__file__))
data_xlsx_path: str = os.path.join(script_path, 'price_data.xlsx')


def dict_to_xlsx(src_dict: dict, dst_xlsx_path: str):
    '''
    Function: 把字典写入到xlsx文件中，可用于最终保存或爬虫失败时数据保存
    '''
    wb = Workbook()
    for k,v in src_dict.items():
        temp_sht = wb.create_sheet(k)
        for j in range(2,14):
            temp_sht.cell(1,j).value = str(j-1)
        i: int = 1
        for kk,vv in v.items():
            i += 1
            temp_sht.cell(i, 1).value = kk
            for kkk,vvv in vv.items():
                temp_sht.cell(i, int(kkk) + 1).value = vvv

    wb.save(dst_xlsx_path)


def get_city_region_dict(city_name: str) -> dict:
    '''
    Function: 获取指定城市下面的下属区划的名称及其对应的房价网址字典
    '''
    # 构造爬虫解析数据，附带headers骗过网站爬虫检测
    city_url: str = 'https://www.anjuke.com/fangjia/' + city_name
    request_header: dict = {"User-Agent": 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}
    r = requests.get(city_url, headers=request_header)
    soup = BeautifulSoup(r.text, 'lxml')

    # 数据解析
    dst_content = soup.find('span', class_='elem-l').find_all('a')
    region_url_dict: dict = {'全部': city_url}
    for a in dst_content:
        region_url_dict[a.string] = str(a['href'])

    # print(region_url_dict)

    return region_url_dict


def get_month_price_from_url(region_url: str) -> dict:
    '''
    Function: 从网址中爬取网址数据
    '''
    # 构造爬虫获取数据
    # 随机选择一个代理头，避免爬虫封禁
    user_agent_list = [
        'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
        'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50',
        'Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0;', 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)', 'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1',
        'Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1',
        'Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11',
        'Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Maxthon 2.0)', 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; TencentTraveler 4.0)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; The World)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Trident/4.0; SE 2.X MetaSr 1.0; SE 2.X MetaSr 1.0; .NET CLR 2.0.50727; SE 2.X MetaSr 1.0)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; 360SE)',
        'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; Avant Browser)',
        'Mozilla/5.0 (hp-tablet; Linux; hpwOS/3.0.0; U; en-US) AppleWebKit/534.6 (KHTML, like Gecko) wOSBrowser/233.70 Safari/534.6 TouchPad/1.0',
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1",
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3",
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24",
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/535.24",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.133 Safari/537.36"]
    request_header: dict = {"User-Agent": random.choice(user_agent_list)}
    r = requests.get(region_url, headers=request_header)
    soup = BeautifulSoup(r.text, 'lxml')

    # 解析房价数据
    month_price_dict: dict = {}
    dst_content = soup.find('div', class_='fjlist-box boxstyle2').find_all('li')
    for li in dst_content:
        y_m_str: str = li.find('b').string
        month: str = re.findall('年(.*)月', y_m_str)[0]
        price: str = li.find('span').string.replace('元/㎡','')
        month_price_dict[month] = price

    return month_price_dict


def get_city_history_house_price(city_name: str, start_year: int, end_year: int) -> dict:
    '''
    Function: 获取某个城市的的从起止年份的房价数据
    '''
    # 请求数据并保存到数据字典中
    region_url_dict: dict = get_city_region_dict(city_name)
    region_history_price_dict: dict = {}

    for k,v in region_url_dict.items():
        region_history_price_dict[k] = {}
        for year in range(start_year, end_year + 1):
            now_time = str(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))    # 获取当前时间
            print(f'【{now_time}】正在爬取：{k}  {year}年 的房价数据')                 # 显示当前爬取的数据
            dst_url: str = v.replace('/' + city_name + '/', '/' + city_name + str(year) + '/')      # 避免出现beijingzhoubian这种，替换后网站出现错误
            # 设置延时避免请求频繁
            DOWNLOAD_DELAY = random.choice([3, 4, 5, 6])
            time.sleep(DOWNLOAD_DELAY)
            try:
                region_history_price_dict[k][year] = get_month_price_from_url(dst_url)
            except:
                region_history_price_dict[k][year] = {}
                dict_to_xlsx(region_history_price_dict, data_xlsx_path)
                print('爬虫异常，目前数据已经临时保存到xlsx文件中！')
                sys.exit()


    return region_history_price_dict


price_dict: dict = get_city_history_house_price(city_name, start_year, end_year)
dict_to_xlsx(price_dict, data_xlsx_path)
print('数据爬取完毕！')
